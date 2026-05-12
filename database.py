import sqlite3
import os
from datetime import datetime

DATABASE = 'invoices.db'


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_type TEXT,
            invoice_title TEXT,
            invoice_code TEXT,
            invoice_number TEXT,
            invoice_date TEXT,
            verify_code TEXT,
            password_area TEXT,
            buyer_name TEXT,
            buyer_tax_id TEXT,
            buyer_address_phone TEXT,
            buyer_account TEXT,
            seller_name TEXT,
            seller_tax_id TEXT,
            seller_address_phone TEXT,
            item_category TEXT,
            item_details TEXT,
            amount REAL,
            tax_amount REAL,
            total_amount REAL,
            total_amount_cn TEXT,
            remarks TEXT,
            payee TEXT,
            reviewer TEXT,
            drawer TEXT,
            import_time TEXT,
            reimburse_category TEXT,
            reimburse_person TEXT,
            reimburse_remark TEXT,
            reimburse_month TEXT,
            pdf_path TEXT,
            is_duplicate INTEGER DEFAULT 0,
            buyer_name_mismatch INTEGER DEFAULT 0,
            buyer_tax_id_mismatch INTEGER DEFAULT 0
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            created_at TEXT
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            tax_id TEXT,
            created_at TEXT
        )
    ''')

    conn.commit()
    conn.close()


def add_invoice(data):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO invoices (
            invoice_type, invoice_title, invoice_code, invoice_number, invoice_date,
            verify_code, password_area, buyer_name, buyer_tax_id, buyer_address_phone,
            buyer_account, seller_name, seller_tax_id, seller_address_phone,
            item_category, item_details, amount, tax_amount, total_amount, total_amount_cn,
            remarks, payee, reviewer, drawer, import_time, pdf_path
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('invoice_type'), data.get('invoice_title'), data.get('invoice_code'),
        data.get('invoice_number'), data.get('invoice_date'), data.get('verify_code'),
        data.get('password_area'), data.get('buyer_name'), data.get('buyer_tax_id'),
        data.get('buyer_address_phone'), data.get('buyer_account'), data.get('seller_name'),
        data.get('seller_tax_id'), data.get('seller_address_phone'), data.get('item_category'),
        data.get('item_details'), data.get('amount'), data.get('tax_amount'),
        data.get('total_amount'), data.get('total_amount_cn'), data.get('remarks'),
        data.get('payee'), data.get('reviewer'), data.get('drawer'),
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'), data.get('pdf_path')
    ))
    invoice_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return invoice_id


def get_all_invoices(filters=None, sort_by=None, sort_order='desc'):
    conn = get_db()
    cursor = conn.cursor()
    query = 'SELECT * FROM invoices WHERE 1=1'
    params = []

    if filters:
        if filters.get('reimburse_month'):
            query += ' AND reimburse_month = ?'
            params.append(filters['reimburse_month'])
        if filters.get('reimburse_person'):
            query += ' AND reimburse_person LIKE ?'
            params.append(f"%{filters['reimburse_person']}%")
        if filters.get('invoice_type'):
            query += ' AND invoice_type = ?'
            params.append(filters['invoice_type'])
        if filters.get('buyer_name'):
            query += ' AND buyer_name LIKE ?'
            params.append(f"%{filters['buyer_name']}%")
        if filters.get('invoice_number'):
            query += ' AND invoice_number LIKE ?'
            params.append(f"%{filters['invoice_number']}%")

    if sort_by:
        query += f' ORDER BY {sort_by} {sort_order}'
    else:
        query += ' ORDER BY import_time DESC'

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]


def get_invoice(invoice_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM invoices WHERE id = ?', (invoice_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def update_invoice(invoice_id, data):
    conn = get_db()
    cursor = conn.cursor()
    fields = []
    params = []
    for key in ['reimburse_category', 'reimburse_person', 'reimburse_remark', 'reimburse_month']:
        if key in data:
            fields.append(f'{key} = ?')
            params.append(data[key])
    if fields:
        params.append(invoice_id)
        cursor.execute(f"UPDATE invoices SET {', '.join(fields)} WHERE id = ?", params)
        conn.commit()
    conn.close()


def delete_invoice(invoice_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT pdf_path FROM invoices WHERE id = ?', (invoice_id,))
    row = cursor.fetchone()
    if row and row['pdf_path'] and os.path.exists(row['pdf_path']):
        os.remove(row['pdf_path'])
    cursor.execute('DELETE FROM invoices WHERE id = ?', (invoice_id,))
    conn.commit()
    conn.close()


def batch_delete_invoices(ids):
    conn = get_db()
    cursor = conn.cursor()
    for invoice_id in ids:
        cursor.execute('SELECT pdf_path FROM invoices WHERE id = ?', (invoice_id,))
        row = cursor.fetchone()
        if row and row['pdf_path'] and os.path.exists(row['pdf_path']):
            os.remove(row['pdf_path'])
        cursor.execute('DELETE FROM invoices WHERE id = ?', (invoice_id,))
    conn.commit()
    conn.close()


def update_invoice_flags(invoice_id, **kwargs):
    conn = get_db()
    cursor = conn.cursor()
    fields = []
    params = []
    for key, value in kwargs.items():
        fields.append(f'{key} = ?')
        params.append(value)
    if fields:
        params.append(invoice_id)
        cursor.execute(f"UPDATE invoices SET {', '.join(fields)} WHERE id = ?", params)
        conn.commit()
    conn.close()


def get_settings():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM settings')
    rows = cursor.fetchall()
    conn.close()
    return {row['key']: row['value'] for row in rows}


def set_setting(key, value):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (key, value))
    conn.commit()
    conn.close()


def get_categories():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT name FROM categories ORDER BY name')
    rows = cursor.fetchall()
    conn.close()
    return [row['name'] for row in rows]


def add_category(name):
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO categories (name, created_at) VALUES (?, ?)',
                       (name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    conn.close()


def get_companies():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM companies ORDER BY id')
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]


def add_company(name, tax_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('INSERT INTO companies (name, tax_id, created_at) VALUES (?, ?, ?)',
                   (name, tax_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    company_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return company_id


def delete_company(company_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM companies WHERE id = ?', (company_id,))
    conn.commit()
    conn.close()


def update_company(company_id, name, tax_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE companies SET name = ?, tax_id = ? WHERE id = ?',
                   (name, tax_id, company_id))
    conn.commit()
    conn.close()


def get_duplicate_invoices():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT invoice_number, COUNT(*) as cnt FROM invoices
        WHERE invoice_number IS NOT NULL AND invoice_number != ''
        GROUP BY invoice_number HAVING cnt > 1
    ''')
    rows = cursor.fetchall()
    conn.close()
    return [row['invoice_number'] for row in rows]


def export_to_excel(filepath, invoice_ids=None):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if invoice_ids:
        conn = get_db()
        cursor = conn.cursor()
        placeholders = ','.join('?' * len(invoice_ids))
        cursor.execute(f'SELECT * FROM invoices WHERE id IN ({placeholders})', invoice_ids)
        rows = cursor.fetchall()
        conn.close()
        data_rows = [dict(row) for row in rows]
    else:
        data_rows = get_all_invoices()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '发票数据'

    headers = [
        '发票类型', '发票标题', '发票代码', '发票号码', '开票日期', '检验码', '密码区',
        '购买方名称', '购买方税号', '购买方地址/电话', '购买方账号',
        '销售方名称', '销售方税号', '销售方地址/电话',
        '开票项目', '开票明细', '开票金额', '开票税额', '价税合计', '价税合计大写',
        '备注', '收款人', '复核人', '开票人', '导入时间',
        '报销分类', '报销人', '我的报销备注', '报销月'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    field_keys = [
        'invoice_type', 'invoice_title', 'invoice_code', 'invoice_number', 'invoice_date',
        'verify_code', 'password_area', 'buyer_name', 'buyer_tax_id', 'buyer_address_phone',
        'buyer_account', 'seller_name', 'seller_tax_id', 'seller_address_phone',
        'item_category', 'item_details', 'amount', 'tax_amount', 'total_amount', 'total_amount_cn',
        'remarks', 'payee', 'reviewer', 'drawer', 'import_time',
        'reimburse_category', 'reimburse_person', 'reimburse_remark', 'reimburse_month'
    ]

    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, key in enumerate(field_keys, 1):
            value = row_data.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else '')
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filepath)
    return filepath


def backup_database(backup_path):
    import shutil
    shutil.copy2(DATABASE, backup_path)
    return backup_path
